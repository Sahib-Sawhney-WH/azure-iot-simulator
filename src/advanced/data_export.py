"""Data export functionality"""
import json
import csv
from datetime import datetime
from typing import Dict, List, Any, Optional
from pathlib import Path

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

from core.logging_config import get_logger

logger = get_logger('data_export')

class DataExporter:
    """Data export functionality for metrics and device data"""
    
    def __init__(self, export_dir: str = "exports"):
        self.export_dir = Path(export_dir)
        self.export_dir.mkdir(exist_ok=True)
    
    def export_metrics_json(self, metrics_data: Dict[str, Any], filename: Optional[str] = None) -> str:
        """Export metrics to JSON format"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"metrics_export_{timestamp}.json"
        
        filepath = self.export_dir / filename
        
        try:
            with open(filepath, 'w') as f:
                json.dump(metrics_data, f, indent=2, default=str)
            
            logger.info(f"Metrics exported to JSON: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error exporting to JSON: {e}")
            raise
    
    def export_metrics_csv(self, metrics_data: Dict[str, Any], filename: Optional[str] = None) -> str:
        """Export metrics to CSV format"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"metrics_export_{timestamp}.csv"
        
        filepath = self.export_dir / filename
        
        try:
            with open(filepath, 'w', newline='') as f:
                writer = csv.writer(f)
                
                # Write system metrics
                writer.writerow(["System Metrics"])
                writer.writerow(["Metric", "Value"])
                
                system_metrics = metrics_data.get("system_metrics", {})
                for key, value in system_metrics.items():
                    writer.writerow([key, value])
                
                writer.writerow([])  # Empty row
                
                # Write device metrics
                writer.writerow(["Device Metrics"])
                writer.writerow(["Device ID", "Messages", "Errors", "Status", "Last Message"])
                
                device_metrics = metrics_data.get("device_metrics", {})
                for device_id, metrics in device_metrics.items():
                    writer.writerow([
                        device_id,
                        metrics.get("message_count", 0),
                        metrics.get("error_count", 0),
                        metrics.get("connection_status", "unknown"),
                        metrics.get("last_message_time", "never")
                    ])
            
            logger.info(f"Metrics exported to CSV: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error exporting to CSV: {e}")
            raise
    
    def export_metrics_excel(self, metrics_data: Dict[str, Any], filename: Optional[str] = None) -> str:
        """Export metrics to Excel format"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl not available - Excel export disabled")
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"metrics_export_{timestamp}.xlsx"
        
        filepath = self.export_dir / filename
        
        try:
            workbook = openpyxl.Workbook()
            
            # System metrics sheet
            sys_sheet = workbook.active
            sys_sheet.title = "System Metrics"
            
            sys_sheet.append(["Metric", "Value"])
            system_metrics = metrics_data.get("system_metrics", {})
            for key, value in system_metrics.items():
                sys_sheet.append([key, value])
            
            # Device metrics sheet
            dev_sheet = workbook.create_sheet("Device Metrics")
            dev_sheet.append(["Device ID", "Messages", "Errors", "Status", "Last Message"])
            
            device_metrics = metrics_data.get("device_metrics", {})
            for device_id, metrics in device_metrics.items():
                dev_sheet.append([
                    device_id,
                    metrics.get("message_count", 0),
                    metrics.get("error_count", 0),
                    metrics.get("connection_status", "unknown"),
                    metrics.get("last_message_time", "never")
                ])
            
            # Historical data sheet
            hist_sheet = workbook.create_sheet("Historical Data")
            hist_sheet.append(["Timestamp", "Messages", "Errors"])
            
            historical_data = metrics_data.get("historical_data", {})
            messages = historical_data.get("messages", [])
            errors = historical_data.get("errors", [])
            
            # Combine historical data
            for msg_data in messages:
                timestamp = msg_data.get("timestamp", "")
                count = msg_data.get("count", 0)
                
                # Find corresponding error data
                error_count = 0
                for err_data in errors:
                    if err_data.get("timestamp") == timestamp:
                        error_count = err_data.get("count", 0)
                        break
                
                hist_sheet.append([timestamp, count, error_count])
            
            workbook.save(filepath)
            
            logger.info(f"Metrics exported to Excel: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            raise
    
    def export_metrics_pdf(self, metrics_data: Dict[str, Any], filename: Optional[str] = None) -> str:
        """Export metrics to PDF format"""
        if not PDF_AVAILABLE:
            raise ImportError("reportlab not available - PDF export disabled")
        
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"metrics_report_{timestamp}.pdf"
        
        filepath = self.export_dir / filename
        
        try:
            doc = SimpleDocTemplate(str(filepath), pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title = Paragraph("Azure IoT Hub Device Simulator - Metrics Report", styles['Title'])
            story.append(title)
            story.append(Spacer(1, 12))
            
            # Export info
            export_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            info = Paragraph(f"Generated: {export_time}", styles['Normal'])
            story.append(info)
            story.append(Spacer(1, 12))
            
            # System metrics
            story.append(Paragraph("System Metrics", styles['Heading2']))
            
            system_metrics = metrics_data.get("system_metrics", {})
            sys_data = [["Metric", "Value"]]
            for key, value in system_metrics.items():
                sys_data.append([key.replace('_', ' ').title(), str(value)])
            
            sys_table = Table(sys_data)
            sys_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(sys_table)
            story.append(Spacer(1, 12))
            
            # Device metrics
            story.append(Paragraph("Device Metrics", styles['Heading2']))
            
            device_metrics = metrics_data.get("device_metrics", {})
            if device_metrics:
                dev_data = [["Device ID", "Messages", "Errors", "Status"]]
                for device_id, metrics in list(device_metrics.items())[:20]:  # Limit to first 20
                    dev_data.append([
                        device_id,
                        str(metrics.get("message_count", 0)),
                        str(metrics.get("error_count", 0)),
                        metrics.get("connection_status", "unknown")
                    ])
                
                dev_table = Table(dev_data)
                dev_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(dev_table)
                
                if len(device_metrics) > 20:
                    note = Paragraph(f"Note: Showing first 20 of {len(device_metrics)} devices", styles['Italic'])
                    story.append(note)
            else:
                story.append(Paragraph("No device metrics available", styles['Normal']))
            
            doc.build(story)
            
            logger.info(f"Metrics exported to PDF: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error exporting to PDF: {e}")
            raise
    
    def export_metrics_html(self, metrics_data: Dict[str, Any], filename: Optional[str] = None) -> str:
        """Export metrics to HTML format"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"metrics_report_{timestamp}.html"
        
        filepath = self.export_dir / filename
        
        try:
            html_content = self._generate_html_report(metrics_data)
            
            with open(filepath, 'w') as f:
                f.write(html_content)
            
            logger.info(f"Metrics exported to HTML: {filepath}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error exporting to HTML: {e}")
            raise
    
    def _generate_html_report(self, metrics_data: Dict[str, Any]) -> str:
        """Generate HTML report content"""
        export_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Azure IoT Hub Device Simulator - Metrics Report</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #2196F3; }}
                h2 {{ color: #1976D2; border-bottom: 2px solid #1976D2; }}
                table {{ border-collapse: collapse; width: 100%; margin: 10px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; font-weight: bold; }}
                .metric-value {{ font-weight: bold; color: #2196F3; }}
                .export-info {{ color: #666; font-size: 12px; }}
            </style>
        </head>
        <body>
            <h1>Azure IoT Hub Device Simulator - Metrics Report</h1>
            <p class="export-info">Generated: {export_time}</p>
            
            <h2>System Metrics</h2>
            <table>
                <tr><th>Metric</th><th>Value</th></tr>
        """
        
        # System metrics
        system_metrics = metrics_data.get("system_metrics", {})
        for key, value in system_metrics.items():
            metric_name = key.replace('_', ' ').title()
            html += f"<tr><td>{metric_name}</td><td class='metric-value'>{value}</td></tr>"
        
        html += """
            </table>
            
            <h2>Device Metrics</h2>
            <table>
                <tr><th>Device ID</th><th>Messages</th><th>Errors</th><th>Status</th><th>Last Message</th></tr>
        """
        
        # Device metrics
        device_metrics = metrics_data.get("device_metrics", {})
        for device_id, metrics in device_metrics.items():
            html += f"""
                <tr>
                    <td>{device_id}</td>
                    <td class='metric-value'>{metrics.get('message_count', 0)}</td>
                    <td class='metric-value'>{metrics.get('error_count', 0)}</td>
                    <td>{metrics.get('connection_status', 'unknown')}</td>
                    <td>{metrics.get('last_message_time', 'never')}</td>
                </tr>
            """
        
        html += """
            </table>
        </body>
        </html>
        """
        
        return html
    
    def get_supported_formats(self) -> List[str]:
        """Get list of supported export formats"""
        formats = ["json", "csv", "html"]
        
        if EXCEL_AVAILABLE:
            formats.append("excel")
        
        if PDF_AVAILABLE:
            formats.append("pdf")
        
        return formats
    
    def export_metrics(self, metrics_data: Dict[str, Any], format: str, filename: Optional[str] = None) -> str:
        """Export metrics in specified format"""
        format = format.lower()
        
        if format == "json":
            return self.export_metrics_json(metrics_data, filename)
        elif format == "csv":
            return self.export_metrics_csv(metrics_data, filename)
        elif format == "excel" and EXCEL_AVAILABLE:
            return self.export_metrics_excel(metrics_data, filename)
        elif format == "pdf" and PDF_AVAILABLE:
            return self.export_metrics_pdf(metrics_data, filename)
        elif format == "html":
            return self.export_metrics_html(metrics_data, filename)
        else:
            raise ValueError(f"Unsupported export format: {format}")
